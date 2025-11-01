import requests
import os
import pdfplumber
import json
import re
import threading
from openpyxl import Workbook, load_workbook
from easyofd import OFD


class RecToForm:
    def __init__(self, in_path, out_path, message, api_key, base_url, max_threads):
        self.in_path = in_path
        self.out_path = out_path
        self.msg = message
        self.api_key = api_key
        self.base_url = base_url
        self.max_threads = max_threads

        self.pdf_informations = []
        self.pdf_names = []
        self.newfiles_path = []
        self.dataframes = []
        self.all_data = []

    def getfile(self):
        for files_path in os.listdir(self.in_path):
            file_path = os.path.join(self.in_path, files_path)
            file_name, ext = os.path.splitext(files_path)
            if ext == ".ofd":
                ofd = OFD()
                ofd.read(file_path, fmt="path")
                pdf_bytes = ofd.to_pdf()
                newfile = os.path.join(self.in_path, f"{file_name}.pdf")
                with open(newfile, "wb") as f:
                    f.write(pdf_bytes)
                self.newfiles_path.append(newfile)
                ofd.disposal()

        for files_path in os.listdir(self.in_path):
            file_path = os.path.join(self.in_path, files_path)
            file_name, ext = os.path.splitext(files_path)
            if ext == ".pdf":
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        self.pdf_informations.append(text)
                self.pdf_names.append(file_name)

        for p in self.newfiles_path:
            os.remove(p)

    # 请求现使用deepseek官方的api
    def useapi(self, pdf_information):
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        }
        data = {
            "messages": [{"role": "user", "content": self.msg + '\n' + pdf_information}],
            "model": "deepseek-chat",
            "stream": False
        }
        resp = requests.post(self.base_url, json=data, headers=headers)
        result = resp.json()
        content = result["choices"][0]["message"]["content"]
        cleaned = re.sub(r"```(?:python)?\n?", "", content, flags=re.DOTALL).strip()
        return cleaned

    def recognize(self):
        threads = []
        lock = threading.Lock()
        sem = threading.Semaphore(self.max_threads)

        def worker(idx, pdf_info):
            with sem:
                try:
                    content = self.useapi(pdf_info).strip()
                    match = re.search(r"(\{.*?}|\[.*?])", content, re.S)

                    if match:
                        json_str = match.group(1)
                        result = json.loads(json_str)
                    else:
                        print(f"文件{idx + 1} 返回非JSON，跳过。")
                        return
                    with lock:
                        self.dataframes.append((idx, result))
                    print(f"文件{idx + 1} {self.pdf_names[idx]} 已分析完成")
                except Exception as e:
                    print(f"文件{idx + 1} 出错: {e}")
        print("开始分析...")
        for idx, info in enumerate(self.pdf_informations):
            t = threading.Thread(target=worker, args=(idx, info))
            t.start()
            threads.append(t)

        for t in threads:
            t.join()

        self.dataframes = [d for _, d in sorted(self.dataframes, key=lambda x: x[0])]

    def _flatten_results(self):
        """将 self.dataframes 中的数据展开成 list"""
        all_data = []
        for item in self.dataframes:
            if isinstance(item, tuple) and len(item) == 2:
                _, result = item
            else:
                result = item
            if isinstance(result, list):
                all_data.extend(result)
            elif isinstance(result, dict):
                all_data.append(result)
        return all_data

    def fill(self):
        wb = Workbook()
        ws = wb.active
        heads = ["序号", "发票代码", "发票号", "发票金额"]
        for i, h in enumerate(heads, 1):
            ws.cell(1, i, h)

        self.all_data = self._flatten_results()

        idx_row = 2
        for data in self.all_data:
            ws.cell(idx_row, 1, idx_row - 1)
            ws.cell(idx_row, 2, data.get("发票代码", ""))
            ws.cell(idx_row, 3, data.get("发票号码", ""))
            ws.cell(idx_row, 4, data.get("发票金额", data.get("价税合计（小写）", "")))
            idx_row += 1

        wb.save(self.out_path)
        print("普通发票表格已填写完成。\n")



    def fill_template(self):
        # 1) 打开模板（务必是 .xlsx）
        template_path = "模板.xlsx"
        save_path = "模板_自动填写版.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active

        # 2) 在前 20 行内根据表头文字定位列号（支持中英别名）
        header_aliases = {
            "低值材料分类号": [
                "分类分类",
                "(必填项)分类分类",
                "低值材料分类号",
                "分类号",
                "MANAGERATTCODE"
            ],
            "资产名称": ["资产名称", "ASSETNAME"],
            "品牌": ["品牌", "BRAND"],
            "规格型号": ["规格型号", "MODEL", "SPEC"],
            "单位": ["单位", "UNIT"],
            "数量": ["数量", "QUANTITY"],
            "单价": ["单价(元)", "单价", "PRICE"],
            "总价": ["总价(元)", "总价", "TOTALPRICE"],
            "供应商": ["供应商（按发票填写）", "供应商", "SUPPLIERID", "SUPPLIER"],
            "发票编号": ["发票编号", "INVOICENO", "INVOICE NO"],
            "开票日期": ["开票日期", "INVOICEDATE", "INVOICE DATE"],
            "存放地址": ["存放地址", "ADDRESS"],
        }

        def find_header_positions(ws):
            pos = {}
            header_row = None
            # 搜索前 20 行寻找表头
            for r in range(1, 21):
                texts = [str(ws.cell(r, c).value).strip() if ws.cell(r, c).value is not None else "" for c in
                         range(1, ws.max_column + 1)]
                if any(t for t in texts):
                    # 尝试匹配别名
                    for key, aliases in header_aliases.items():
                        for c, t in enumerate(texts, start=1):
                            if any(alias in t for alias in aliases):
                                pos[key] = c
                                header_row = r
                    # 如果已匹配到关键字段，认为本行就是表头
                    if "资产名称" in pos and "发票编号" in pos:
                        break
            return header_row, pos

        header_row, col_pos = find_header_positions(ws)
        if not header_row or "资产名称" not in col_pos:
            raise RuntimeError("未在模板中识别到表头，请确认表头包含“资产名称 / 发票编号”等字段。")

        # 3) 计算起始写入行：从表头下一行开始，找到第一个“资产名称空”的行
        def first_empty_row():
            start = header_row + 1
            c_name = col_pos["资产名称"]
            r = start
            # 连续向下找到第一个该列为空的行
            while True:
                v = ws.cell(r, c_name).value
                if v is None or str(v).strip() == "":
                    return r
                r += 1

        start_row = first_empty_row()

        # 4) 将 self.dataframes 规整成 list[dict]
        self.all_data = self._flatten_results()

        # 5) 工具函数：把字符串里的人民币符号/逗号清掉，保留两位小数
        def to_decimal_str(x):
            if x is None:
                return None
            s = str(x).replace("¥", "").replace(",", "").strip()
            try:
                return f"{float(s):.2f}"
            except Exception as e:
                print(f"[警告] 无法转换为数字：{s}，错误：{e}")
                return s

        # 6) 逐行写入（只写到能识别到列的字段）
        r = start_row
        for data in self.all_data:
            def write_field(field_key, value):
                if field_key in col_pos:
                    ws.cell(r, col_pos[field_key], value)

            write_field("低值材料分类号", data.get("低值材料分类号"))
            write_field("资产名称", data.get("资产名称") or data.get("商品名") or data.get("项目名称"))
            write_field("品牌", data.get("品牌"))
            # 模型会把“规格型号”拆成 MODEL/SPEC；统一写到“规格型号”列
            write_field("规格型号", data.get("规格型号") or data.get("MODEL") or data.get("SPEC"))
            write_field("单位", data.get("单位") or "件")
            write_field("数量", data.get("数量") or "1")

            if "单价" in col_pos:
                ws.cell(r, col_pos["单价"], to_decimal_str(data.get("单价")))
            if "总价" in col_pos:
                ws.cell(r, col_pos["总价"],
                        to_decimal_str(data.get("总价") or data.get("价税合计（小写）") or data.get("价税合计")))

            write_field("供应商", data.get("供应商") or data.get("销售方") or data.get("销售方名称"))
            write_field("发票编号", data.get("发票编号") or data.get("发票号码"))
            write_field("开票日期", data.get("开票日期"))
            write_field("存放地址", data.get("存放地址"))

            r += 1

        wb.save(save_path)
        print(f"模板已填写完成：{save_path}\n")


def choice():
    print("请选择功能模式：")
    print("1. 普通发票信息提取 制作者：Yunxi_Zhu")
    print("2. 大创低值材料资产入库模板自动导入 制作者：kkghrsbsb")
    print()

    while True:
        try:
            mode = int(input("输入数字选择模式："))
            match mode:
                case 1 | 2:
                    print()
                    return mode
                case _:
                    print("输入无效，请输入存在序号")
        except ValueError:
            print("输入无效")


def hint(mode):
    print("(1) 请确保你已经更改代码中的\"api_key\"")
    print("(2) 请确保你已创建“发票”文件夹并放入发票")
    flag = "n"
    match mode:
        case 1:
            flag = input("[y/n]：")

        case 2:
            print("(3)确保文件夹中有模板.xlsx文件")
            flag = input("[y/n]：")

        case _:
            print("未知模式")
            flag = "n"
    print()

    return flag.lower() == "y"


if __name__ == "__main__":
    in_path = "发票"
    out_path = "发票信息.xlsx"

    # 填写个人api密钥
    api_key = "sk-ff4d142d2d744cb0a86d0e572d9b590a"

    # api调用平台
    base_url = "https://api.deepseek.com/chat/completions"

    # 识别过程中使用的最大线程数
    max_threads = 5

    mode = choice()
    match mode:
        case 1:
            msg = "请从以下发票文本中提取发票代码、发票号码、发票金额（价税合计中的小写金额），若缺少数据则令发票代码与发票号码相同，若出现乱码请转换为可读文字，所有字段值为字符串，用双引号包裹，仅返回JSON列表（数组）[{...}, {...}]，不要附加解释。"
            hint(mode)

            rtf = RecToForm(in_path, out_path, msg, api_key, base_url, max_threads)
            rtf.getfile()
            rtf.recognize()
            rtf.fill()

        case 2:
            msg = "请从以下电子发票文本中提取字段：低值材料分类号（根据资产名称推断，低值请填写2或02,材料请填写3或03）、资产名称、品牌、规格型号、单位、数量、单价、总价（价税合计小写）、供应商、发票编号、开票日期（YYYY-MM-DD）、存放地址（根据资产名称推断，十字以内）；所有数值为正数，数量是正整数，字段值均为字符串，用双引号包裹，仅返回JSON列表（数组）[{...}, {...}]，不要附加解释。资产的宽泛存放地址为："
            hint(mode)
            def msg_additional(msg):
                msg_ad = input("为写明材料存放地址，请给你所属学院简称（如智控学院）：")
                print()
                return msg + msg_ad

            msg_additional(msg)

            rtf = RecToForm(in_path, out_path, msg, api_key, base_url, max_threads)
            rtf.getfile()
            rtf.recognize()
            rtf.fill_template()

    print("！！ai识别出的结果小概率可能有误或出现未填上的，注意根据发票编号进行核实！！")
